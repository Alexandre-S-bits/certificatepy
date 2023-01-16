from PIL import ImageDraw, ImageFont
from tkinter import *
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
import PIL.Image
import os

root = Tk()
root.title('Certificatepy')
root.configure(background='#2A2A2C')
root.geometry('750x500+100+100')
cores = {'padrao': '#2A2A2C', 'azul': '#183A84', 'verde': '#04A777', 'fonte_branca': '#ebe4e4',
         'cinza': '#b6b6b6', 'vermelho': '#d64933'}
root.maxsize(850,600)


var_checkbox = IntVar()
var_fonte = IntVar()
nome_fonte = ''


def selecionar_planilha():
    entry_caminho_planilha.delete(0, "end")
    str_planilha = filedialog.askopenfilename(title="Arquivos do Excel", filetypes=(('Excel files', ['*.xlsx', '*.xlsm', '*.xls']),))
    entry_caminho_planilha.insert(0, str_planilha)
    caminho_planilha = entry_caminho_planilha.get()
    planilha = load_workbook(caminho_planilha, data_only=True)  # ler somente os dados da planilha
    nomes_guias = planilha.sheetnames
    combobox_guias.config(values=nomes_guias)


def selecionar_imagem():
    entry_imagem.delete(-1, "end") # Ao chamar a função deleta o conteudo da caixa de texto
    str_imagem = filedialog.askopenfilename(title="Selecione a Imagem",
                                            filetypes=(('Image files', ['*.jpg', '*.png']),))
    entry_imagem.insert(1, str_imagem) # Insere o caminho da imagem na caixa de texto (entry)


def imagem_padrao():
    if var_checkbox.get() == 1:
        entry_imagem.insert(1, 'certificate.png') # Imagem padrao, caso o caixa de seleçao esteja ativada a caixa de texto é desabilatada
        entry_imagem.config(state='disable')
    elif var_checkbox.get() == 0:
        entry_imagem.config(state='normal')
        entry_imagem.delete(0, 'end')


def obter_fonte():
    global nome_fonte
    if var_fonte.get() == 1:
        caminho_fonte = f'{os.getcwd()}/fonts/tipo1.ttf'
        if os.path.isfile == False:
            caminho_fonte = f'{os.getcwd()}/fonts/tipo1.otf'
        nome_fonte = [caminho_fonte, 55]

    elif var_fonte.get() == 2:
        caminho_fonte = f'{os.getcwd()}/fonts/tipo2.ttf'
        if os.path.isfile == False:
            caminho_fonte = f'{os.getcwd()}/fonts/tipo2.otf'
        nome_fonte = [caminho_fonte, 55]

    elif var_fonte.get() == 3:
        caminho_fonte = f'{os.getcwd()}/fonts/tipo3.ttf'
        if os.path.isfile == False:
            caminho_fonte = f'{os.getcwd()}/fonts/tipo3.otf'
        nome_fonte = [caminho_fonte, 55]

    return nome_fonte


def obter_coordenadas_conteudo(caminho_planilha):
    planilha = load_workbook(caminho_planilha, data_only=True)  # ler somente os dados da planilha
    guia_coordenadas = planilha['coordenadas']
    coordenadas = {}
    contador_numero_linha = 1
    while True:
        coordenadas[f"{guia_coordenadas[f'A{contador_numero_linha}'].value}"] = f"[{guia_coordenadas[f'C{contador_numero_linha}'].value}, {guia_coordenadas[f'C{contador_numero_linha+1}'].value}]"
        contador_numero_linha += 2
    
    return coordenadas


def gerar_certificado():
    contador_numero_linha = 2
    nome_fonte = obter_fonte()
    caminho_planilha = entry_caminho_planilha.get()
    guia = combobox_guias.get()
    caminho_template = entry_imagem.get()
    planilha = load_workbook(caminho_planilha, data_only=True)  # ler somente os dados da planilha
    nome_guia = planilha[guia]  # abre na guia especificada dentro dos colchetes"""
    
    # Obtem coordenadas aonde "desenhar" o conteudo na imagem
    dicionario_coordenadas = obter_coordenadas_conteudo(caminho_planilha)
    lista_chaves_dicionario = list(dicionario_coordenadas.keys())

    # Carregar uma fonte de um arquivo TTF
    fonte_atual = ImageFont.truetype(nome_fonte[0], nome_fonte[1])
    informacoes_vazias = []
    while True:
        nome = nome_guia[f'A{contador_numero_linha}'].value

        if nome == ['', '-'] or nome is None:
            break

        curso = nome_guia[f'B{contador_numero_linha}'].value
        aproveitamento = str(nome_guia[f'C{contador_numero_linha}'].value)
        carga_horaria = nome_guia[f'D{contador_numero_linha}'].value
        data_inicial = nome_guia[f'E{contador_numero_linha}'].value
        data_conclusao = nome_guia[f'F{contador_numero_linha}'].value
        dia_emissao = nome_guia[f'I{contador_numero_linha}'].value
        mes_emissao = nome_guia[f'J{contador_numero_linha}'].value
        ano_emissao = nome_guia[f'K{contador_numero_linha}'].value

        # Caminho do Template Base
        template = PIL.Image.open(caminho_template, mode='r')

        # criar um canvas no topo da imagem
        draw = ImageDraw.Draw(template)
        draw.text((dicionario_coordenadas[lista_chaves_dicionario[0]][0], dicionario_coordenadas[lista_chaves_dicionario[0]][1]), nome, fill=(0, 0, 0), anchor='ms', font=fonte_atual)
        if len(curso) >= 40:
            draw.text((dicionario_coordenadas[lista_chaves_dicionario[0]][0], dicionario_coordenadas[lista_chaves_dicionario[0]][1]), curso, fill=(0, 0, 0), anchor='l', font=(nome_fonte[0], 40))
        elif len(curso) < 40:
            draw.text((dicionario_coordenadas[lista_chaves_dicionario[0]][0], dicionario_coordenadas[lista_chaves_dicionario[0]][1]), curso, fill=(0, 0, 0), anchor='ms', font=fonte_atual)
        draw.text((dicionario_coordenadas[lista_chaves_dicionario[0]][0], dicionario_coordenadas[lista_chaves_dicionario[0]][1]), aproveitamento, fill=(0, 0, 0), anchor='ms', font=fonte_atual)
        draw.text((dicionario_coordenadas[lista_chaves_dicionario[0]][0], dicionario_coordenadas[lista_chaves_dicionario[0]][1]), str(carga_horaria), fill=(0, 0, 0), anchor='ms', font=fonte_atual)
        draw.text((dicionario_coordenadas[lista_chaves_dicionario[0]][0], dicionario_coordenadas[lista_chaves_dicionario[0]][1]), str(data_inicial), fill=(0, 0, 0), anchor='ms', font=fonte_atual)
        draw.text((dicionario_coordenadas[lista_chaves_dicionario[0]][0], dicionario_coordenadas[lista_chaves_dicionario[0]][1]), str(data_conclusao), fill=(0, 0, 0), anchor='ms', font=fonte_atual)
        draw.text((dicionario_coordenadas[lista_chaves_dicionario[0]][0], dicionario_coordenadas[lista_chaves_dicionario[0]][1]), dia_emissao, fill=(0, 0, 0), anchor='ms', font=fonte_atual)
        draw.text((dicionario_coordenadas[lista_chaves_dicionario[0]][0], dicionario_coordenadas[lista_chaves_dicionario[0]][1]), mes_emissao, fill=(0, 0, 0), anchor='ms', font=fonte_atual)
        draw.text((dicionario_coordenadas[lista_chaves_dicionario[0]][0], dicionario_coordenadas[lista_chaves_dicionario[0]][1]), ano_emissao, fill=(0, 0, 0), anchor='ms', font=fonte_atual)

        # Salvar a imagem
        template.save(f'{nome}_{curso}.png')
        contador_numero_linha += 1

    return messagebox.showinfo('Concluido', f'{contador_numero_linha-1} certificado(s) gerado(s)!')


##########################################  INTERFACE  ##########################################
frame_principal = Frame(root)
frame_principal.pack(fill=BOTH, expand=1)
canvas = Canvas(frame_principal, bg=cores['padrao'])
canvas.pack(fill=BOTH, side=LEFT, expand=1)
sb = ttk.Scrollbar(frame_principal, orient='vertical', command=canvas.yview)
sb.pack(side=RIGHT, fill=Y)

canvas.configure(yscrollcommand=sb.set)
canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

frame_canvas = Frame(canvas)
canvas.create_window((0, 0), window=frame_canvas, anchor='nw')

lbtitulo = Label(canvas, text='CertificadosPY', fg=cores["fonte_branca"], bg=cores['azul'], font=('Helvitica', 14))
lbtitulo.pack(side=TOP, ipadx=6, fill=X)

lbespaco_esq = Label(frame_canvas, bg=cores['padrao'])
lbespaco_esq.pack(side=LEFT, fill=Y, ipadx=12)
lbespaco_dir = Label(frame_canvas, bg=cores['padrao'])
lbespaco_dir.pack(side=RIGHT, fill=Y, ipadx=12, anchor="e")

##########################################  FRAME 1  ##########################################
frame1 = Frame(frame_canvas, bg=cores['padrao'])
frame1.pack(side=RIGHT, fill=BOTH, expand=1, ipady=20)
lbespaco_1 = Label(frame1, bg=cores['padrao'])
lbespaco_1.pack(side=LEFT, ipadx=4, fill=Y)

frame1esq = Frame(frame1, bg=cores['padrao'])
frame1esq.pack(side=LEFT, fill=BOTH, expand=1)
lbespacoesq1 = Label(frame1esq, bg=cores['padrao'], height=3)
lbespacoesq1.pack(side=TOP, ipadx=4, fill=X)
btn_caminho_planilha = Button(frame1esq, fg=cores["fonte_branca"], bg=cores["azul"], text="Selecione a planilha",
                              command=selecionar_planilha, font=('Helvitica', 12), activebackground='green')
btn_caminho_planilha.pack(side=TOP, padx=2, anchor='nw')
entry_caminho_planilha = Entry(frame1esq, bg=cores['cinza'], font=('Helvitica', 12), width=40)
entry_caminho_planilha.pack(side=TOP, anchor='sw')

lbespaco_2 = Label(frame1, bg=cores['padrao'])
lbespaco_2.pack(side=LEFT, ipadx=4, fill=Y)

lbespacoesq2 = Label(frame1esq, bg=cores['padrao'], height=1)
lbespacoesq2.pack(side=TOP, ipadx=4, fill=X)
lb_guias = Label(frame1esq, text=" Nome da guia ", fg=cores["fonte_branca"], bg=cores['azul'], font=('Helvitica', 12))
lb_guias.pack(side=TOP, padx=4, anchor='w')
combobox_guias = ttk.Combobox(frame1esq)
combobox_guias.pack(side=TOP, padx=2, anchor='w')

##########################################  FRAME 2  ##########################################
frame2 = Frame(frame_canvas, bg=cores['padrao'])
frame2.pack(side=BOTTOM, fill=BOTH, expand=1, ipady=20, before=frame1)
lbespaco_1 = Label(frame2, bg=cores['padrao'])
lbespaco_1.pack(side=LEFT, fill=Y, ipadx=4)

frame2esq = Frame(frame2, bg=cores['padrao'])
frame2esq.pack(side=LEFT, fill=BOTH)
btn_imagem = Button(frame2esq, fg=cores["fonte_branca"], bg=cores["azul"], text="Escolher imagem",
                    command=selecionar_imagem, font=('Helvitica', 12))
btn_imagem.pack(side=TOP, padx=2, anchor='nw')
entry_imagem = Entry(frame2esq, bg=cores['cinza'], font=('Helvitica', 12), width=40)
entry_imagem.pack(side=TOP, expand=1, anchor='nw')

checkbox_imagem = Checkbutton(frame2esq, text='Padrão', variable=var_checkbox, command=imagem_padrao,
                              fg=cores['fonte_branca'], selectcolor=cores['padrao'], font=('Helvitica', 10), bg=cores['padrao'])
checkbox_imagem.pack(side=RIGHT, before=entry_imagem, anchor='ne')

lbespaco_2 = Label(frame2, bg=cores['padrao'])
lbespaco_2.pack(side=LEFT, fill=Y, expand=1, ipadx=4)

frame2dir = Frame(frame2, bg=cores['padrao'])
frame2dir.pack(side=LEFT, fill=BOTH, expand=1)
lb_fonte = Label(frame2dir, text=" Selecione a fonte ", fg=cores["fonte_branca"], bg=cores['azul'], font=('Helvitica', 12))
lb_fonte.pack(side=TOP, padx=5, anchor='w')

# Radiobutton 
radiobutton1 = Radiobutton(frame2dir, text='Fonte 1', variable=var_fonte, value=1, bg=cores['padrao'],
                  fg=cores['fonte_branca'], selectcolor=cores['padrao'], font=('Helvitica', 12))
radiobutton1.pack(side=TOP, anchor='w')
radiobutton2 = Radiobutton(frame2dir, text='Fonte 2', variable=var_fonte, value=2, bg=cores['padrao'],
                  fg=cores['fonte_branca'], selectcolor=cores['padrao'], font=('Helvitica', 12))
radiobutton2.pack(side=TOP, anchor='w')
radiobutton3 = Radiobutton(frame2dir, text='Fonte 3', variable=var_fonte, value=3, bg=cores['padrao'],
                  fg=cores['fonte_branca'], selectcolor=cores['padrao'], font=('Helvitica', 12))
radiobutton3.pack(side=TOP, anchor='w')
radiobutton1.select() # Seleciona essa opção por padrão

##########################################  FRAME 3  ##########################################
frame3 = Frame(frame_canvas, bg=cores['padrao'])
frame3.pack(side=BOTTOM, fill=BOTH, expand=1, ipady=40, before=frame2)

lb_x = Label(frame3, text="", fg=cores["fonte_branca"], bg=cores['padrao'], font=('Helvitica', 12))
lb_x.pack(fill=X)

btn_gerar = Button(frame3, fg=cores["fonte_branca"], bg=cores["azul"], text="Gerar Certificados",
                   command=gerar_certificado, font=('Helvitica', 12))
btn_gerar.pack(padx=2, anchor='s')

root.mainloop()
